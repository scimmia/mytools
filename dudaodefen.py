# coding=utf-8
# 复制删除法按地市法人分割汇总文件
import datetime
import os
from tkinter import *
from tkinter.scrolledtext import ScrolledText
from tkinter.ttk import *
from tkinter.filedialog import askopenfilename, askdirectory
import openpyxl
import xlrd
from copy import copy

import _thread
from utils import orgs


def select_folder():
    folder_path = askdirectory()
    pathFolder.set(folder_path)


def select_file():
    file_path = askopenfilename(filetypes=[('XLSX', '*.xlsx'), ('All Files', '*')])
    pathFile.set(file_path)


def show_log(text):
    logs.insert(END, str(text) + "\n")
    logs.update()
    print(text)


def start_work():
    # if len(pathFile.get()) <= 0:
    #     showinfo('提示', '先选择文件')
    #     return
    source_file = pathFile.get()
    source_folder = pathFolder.get()
    files = os.listdir(source_folder)
    print((files))
    datas = list()
    # datas = ['', 0, '']
    # datas = {}
    for file in files:
        full_path = os.path.join(source_folder, file)
        print(full_path)
        if os.path.isfile(full_path):
            result = get_xls_data(full_path)
            # for i in range(0,3):
            #     datas[i] = datas[i] + result[i]
            datas.append(result)
            # org, result = get_xls_data(full_path)
            # datas[org] = result
    result = list()
    for i in range(4, 17):
        result.append(['', 0, ''])
    for data in datas:
        for i in range(0, len(data)):
            result[i][0] += data[i][0]
            result[i][1] += data[i][1]
            result[i][2] += data[i][2]
    for i in range(0, len(result)):
        result[i][1] = result[i][1] / float(len(datas))
    print(result)
    write_city_file(result)
    return


def get_xls_data(full_path):
    # org = ''
    datas = list()
    try:
        book = xlrd.open_workbook(full_path)
        sh = book.sheet_by_name('督导打分表')
        org = sh.cell(2, 0).value.replace('被督导单位：', '')
        print(org)
        for i in range(4, 17):
            cell_a = sh.cell(i, 6)
            cell_b = sh.cell(i, 7)
            cell_c = sh.cell(i, 8)
            data = ['', 0, '']
            if cell_a.ctype == 1 and len(cell_a.value) > 0:
                data[0] = org + cell_a.value + '\n'
            elif cell_a.ctype == 2:
                data[0] = org + str(int(cell_a.value)) + '\n'
            if cell_b.ctype == 2:
                data[1] = int(cell_b.value)
            if cell_c.ctype == 1 and len(cell_c.value) > 0:
                data[2] = org + cell_c.value + '\n'
            elif cell_c.ctype == 2:
                data[2] = org + str(int(cell_c.value)) + '\n'
            # data[2] = data[2] + str(sh.cell(i, 8).value) + '\n'
            # data.append(org + sh.cell(i, 6).value)
            # data.append(sh.cell(i, 7).value)
            # data.append(org + sh.cell(i, 8).value)
            datas.append(data)
    except Exception as e:
        show_log('出错了')
        show_log(str(e))
    return datas


def write_city_file(datas):
    try:
        book = openpyxl.load_workbook('dudao_template.xlsx')
        sheet = book.active
        for i in range(0, len(datas)):
            sheet.cell(5 + i, 7).value = datas[i][0]
            sheet.cell(5 + i, 8).value = datas[i][1]
            sheet.cell(5 + i, 9).value = datas[i][2]

        source_folder = pathFolder.get()
        last_sp = source_folder.rfind('/')
        org = source_folder[last_sp + 1:]
        book.save('dudao_out/%s-%s农商银行督导打分表（2021-2季度）.xlsx' % (orgs[org], org))
        book.close()
        show_log('%s-%s已完成' % (orgs[org], org))
    except Exception as e:
        show_log(str(e))
    return


def begin_work():
    _thread.start_new_thread(start_work, ())


def start_merge_work():
    source_folder = pathFolder.get()
    files = os.listdir(source_folder)
    book = openpyxl.load_workbook('dudao_template_summary.xlsx')
    sheet = book['总结']
    current_row = 4
    for file in files:
        full_path = os.path.join(source_folder, file)
        print(full_path)
        if os.path.isfile(full_path):
            org = full_path.replace(pathFolder.get(), '')[1:6]
            print(org)
            if full_path.find('.xlsx') > 0:
                result, data_scores = get_xlsx_data_with_total(full_path)
            else:
                result, data_scores = get_xls_data_with_total(full_path)
            for i in range(0, len(data_scores)):
                _ = sheet.cell(column=1 + i, row=current_row, value=data_scores[i])
            current_row += 1
    book.save('dudao_out/dudao_summary.xlsx')
    book.close()
    show_log('已完成')
    return


def get_xls_data_with_total(full_path):
    # org = ''
    datas = list()
    data_scores = list()
    try:
        book = xlrd.open_workbook(full_path)
        sh = book.sheet_by_index(0)
        names = book.sheet_names()
        for name in names:
            if name.find('打分') >= 0:
                sh = book.sheet_by_name(name)
                break
        org = full_path.replace(pathFolder.get(), '')[1:6]
        data_scores.append(org[0:2])
        data_scores.append(org)
        print(org)
        total = 0
        for i in range(4, 17):
            cell_a = sh.cell(i, 6).value
            cell_b = sh.cell(i, 7).value
            cell_c = sh.cell(i, 8).value
            score = 0
            # data = ['', 0, '']
            # if isinstance(cell_a, str) and len(cell_a) > 0:
            #     data[0] = org + cell_a + '\n'
            # elif isinstance(cell_a, (int, float)):
            #     data[0] = org + str(int(cell_a)) + '\n'
            if isinstance(cell_b, (int, float)):
                score = int(cell_b)
                # data[1] = score
                data_scores.append(score)
            else:
                data_scores.append('')
            # if isinstance(cell_c, str) and len(cell_c) > 0:
            #     data[2] = org + cell_c + '\n'
            # elif isinstance(cell_c, (int, float)):
            #     data[2] = org + str(int(cell_c)) + '\n'
            datas.append([cell_a, cell_b, cell_c])
            total += score
        datas.append(['', total, ''])
        data_scores.append(total)

    except Exception as e:
        show_log('%s出错了' % full_path)
        show_log(str(e))
    return datas, data_scores


def get_xlsx_data_with_total(full_path):
    # org = ''
    datas = list()
    data_scores = list()
    try:
        book = openpyxl.load_workbook(full_path)
        sh = book.active
        names = book.get_sheet_names()
        for name in names:
            if name.find('打分') >= 0:
                sh = book[name]
                break
        org = full_path.replace(pathFolder.get(), '')[1:6]
        data_scores.append(org[0:2])
        data_scores.append(org)
        print(org)
        total = 0
        for i in range(5, 18):
            score = 0
            cell_a = sh.cell(i, 7).value
            cell_b = sh.cell(i, 8).value
            cell_c = sh.cell(i, 9).value
            datas.append([cell_a, cell_b, cell_c])
            # data = ['', 0, '']
            # if isinstance(cell_a, str) and len(cell_a) > 0:
            #     data[0] = org + cell_a + '\n'
            # elif isinstance(cell_a, (int, float)):
            #     data[0] = org + str(int(cell_a)) + '\n'
            if isinstance(cell_b, (int, float)):
                score = int(cell_b)
                # data[1] = score
                data_scores.append(score)
            else:
                data_scores.append('')
            # if isinstance(cell_c, str) and len(cell_c) > 0:
            #     data[2] = org + cell_c + '\n'
            # elif isinstance(cell_c, (int, float)):
            #     data[2] = org + str(int(cell_c)) + '\n'
            # datas.append(data)
            total += score
        datas.append(['', total, ''])
        data_scores.append(total)

    except Exception as e:
        show_log('%s出错了' % full_path)
        show_log(str(e))
    return datas, data_scores


def begin_merge_work():
    _thread.start_new_thread(start_merge_work, ())


def start_copy_work():
    source_folder = pathFolder.get()
    files = os.listdir(source_folder)
    wb = openpyxl.load_workbook('dudao_template.xlsx')
    ws = wb['督导打分表']

    book = openpyxl.load_workbook('dudao_template_summary.xlsx')
    sheet = book['督导打分表']
    current_row = 2

    sheet_sum = book['总结']
    current_row_sum = 4

    for file in files:
        full_path = os.path.join(source_folder, file)
        print(full_path)
        if os.path.isfile(full_path):
            org = full_path.replace(pathFolder.get(), '')[1:6]
            show_log(org)
            if full_path.find('.xlsx') > 0:
                result, data_scores = get_xlsx_data_with_total(full_path)
            else:
                result, data_scores = get_xls_data_with_total(full_path)
            write_template(ws, sheet, current_row, org)
            for i in range(0, len(result)):
                _ = sheet.cell(column=7, row=current_row + i, value=result[i][0])
                _ = sheet.cell(column=8, row=current_row + i, value=result[i][1])
                _ = sheet.cell(column=9, row=current_row + i, value=result[i][2])
                # _ = sheet.cell(column=3 + i, row=current_row, value=data_scores[i])
            current_row += 15

            for i in range(0, len(data_scores)):
                _ = sheet_sum.cell(column=1 + i, row=current_row_sum, value=data_scores[i])
            current_row_sum += 1

    book.save('dudao_out/dudao_all.xlsx')
    book.close()
    wb.close()
    show_log('已完成')

    show_log('已完成')
    return


def begin_copy_work():
    _thread.start_new_thread(start_copy_work, ())


def get_template_data():
    wb = openpyxl.load_workbook('dudao_template.xlsx')
    ws = wb['督导打分表']
    result = list()
    max_row = ws.max_row  # 最大行数
    max_column = ws.max_column  # 最大列数
    for m in range(1, max_row + 1):
        result_row = list()
        for n in range(1, max_column + 1):
            result_row.append(ws.cell(m, n).value)
        result.append(result_row)
    return result


def write_template(source_ws, work_sheet, start_rows, org):
    max_row = source_ws.max_row  # 最大行数
    max_column = source_ws.max_column  # 最大列数
    for m in range(2, max_row + 1):
        for n in range(1, max_column + 1):
            cell_from = source_ws.cell(m, n)
            cell_to = work_sheet.cell(start_rows - 2 + m, n)
            cell_to.value = cell_from.value
            if cell_from.has_style:
                cell_to.font = copy(cell_from.font)
                cell_to.border = copy(cell_from.border)
                cell_to.fill = copy(cell_from.fill)
                cell_to.number_format = copy(cell_from.number_format)
                cell_to.protection = copy(cell_from.protection)
                cell_to.alignment = copy(cell_from.alignment)
    for m in range(2, max_row + 1):
        work_sheet.row_dimensions[start_rows - 2 + m].height = source_ws.row_dimensions[m].height
    work_sheet.merge_cells(start_row=start_rows , start_column=2, end_row=start_rows + 2, end_column=2)
    work_sheet.merge_cells(start_row=start_rows + 3, start_column=2, end_row=start_rows + 8, end_column=2)
    work_sheet.cell(start_rows, 1, value=org)
    work_sheet.merge_cells(start_row=start_rows, start_column=1, end_row=start_rows + 12, end_column=1)


def main():
    logs.grid(row=0, column=0, rowspan=6)

    Button(root, text='选择文件夹', command=select_folder).grid(row=0, column=1)
    Entry(root, textvariable=pathFolder).grid(row=1, column=1)
    # Button(root, text='选择模板文件', command=select_file).grid(row=1, column=1)
    # Entry(root, textvariable=pathFile).grid(row=1, column=2)
    # Label(root, text='汇总文件名称', ).grid(row=2, column=1)
    # Entry(root, textvariable=target_file).grid(row=2, column=2)

    Button(root, text='合并文件', command=begin_copy_work).grid(row=3, column=1, columnspan=1)
    Button(root, text='合并汇总文件', command=begin_merge_work).grid(row=4, column=1, columnspan=1)
    Button(root, text='开始', command=begin_work).grid(row=5, column=1, columnspan=1)
    # Label(root, text='需要另存为xlsx格式\n需按摘取字段排序好\n空行过多需先清除内容后删除', ).grid(row=5, column=2, columnspan=1)
    root.mainloop()


root = Tk()
root.wm_title('督导得分汇总')
logs = ScrolledText(root, width=40, height=30)
pathFolder = StringVar()
pathFile = StringVar()
target_file = StringVar()
if __name__ == '__main__':
    main()
