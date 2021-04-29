# coding=utf-8
import datetime
from tkinter import *
from tkinter.scrolledtext import ScrolledText
from tkinter.ttk import *
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import *

import openpyxl
import shutil


def unmerge():
    try:
        sheet_names = []
        sheet_details = {}
        start = start_line.get()
        end = end_line.get()
        col = col_line.get()
        print(start_line.get())
        print(end_line.get())
        print(col_line.get())
        book = openpyxl.load_workbook(pathFile.get())
        sheet = book.get_active_sheet()
        merges = sheet.merged_cell_ranges
        for m in merges:
            if m.min_row >= start and m.max_row <= end:
                sheet.unmerge_cells(m.coord)
        book.save(pathFile.get() + '--unmerge.xlsx')
    except Exception as e:
        showlog(str(e))
    showlog('finished')


# 先将取消合并的单元格，再操作。
def split_file():
    try:
        sheet_names = []
        sheet_details = {}
        start = start_line.get()
        end = end_line.get()
        col = col_line.get()
        print(start_line.get())
        print(end_line.get())
        print(col_line.get())
        file_path = pathFile.get()
        book = openpyxl.load_workbook(file_path)
        sheet = book.get_active_sheet()
        name_temp = None
        start_temp = start
        for i in range(start, end + 1):
            cel = sheet.cell(row=i, column=col).value
            if cel is not None:
                if name_temp is not None:
                    sheet_names.append(name_temp)
                    sheet_details[name_temp] = [start_temp, i - 1]
                name_temp = cel
                start_temp = i
        if name_temp is not None:
            sheet_names.append(name_temp)
            sheet_details[name_temp] = [start_temp, end]
        # book.close()

        for sheet_name in sheet_names:
            try:
                showlog(sheet_name)
                new_file_name = '%s--%s.xlsx' % (file_path, sheet_name)
                # shutil.copyfile(file_path, new_file_name)
                book = openpyxl.load_workbook(file_path)
                sheet = book.get_active_sheet()
                sheet.title = sheet_name
                temp_start = sheet_details[sheet_name][0]
                temp_end = sheet_details[sheet_name][1] + 1
                if temp_end < end:
                    sheet.delete_rows(temp_end, end + 1 - temp_end)
                if temp_start > start:
                    sheet.delete_rows(start, temp_start - start)
                book.save(new_file_name)
            except Exception as e:
                showlog(str(e))

    except Exception as e:
        showlog(str(e))
    showlog('finished')


# 先将取消合并的单元格，再操作。
def doIt():
    try:
        sheet_names = []
        sheet_details = {}
        start = start_line.get()
        end = end_line.get()
        col = col_line.get()
        print(start_line.get())
        print(end_line.get())
        print(col_line.get())
        book = openpyxl.load_workbook(pathFile.get())
        sheet = book.get_active_sheet()
        name_temp = None
        start_temp = start
        for i in range(start, end + 1):
            cel = sheet.cell(row=i, column=col).value
            if cel is not None:
                if name_temp is not None:
                    sheet_names.append(name_temp)
                    sheet_details[name_temp] = [start_temp, i - 1]
                name_temp = cel
                start_temp = i
        if name_temp is not None:
            sheet_names.append(name_temp)
            sheet_details[name_temp] = [start_temp, end]

        for sheet_name in sheet_names:
            copy_sheet = book.copy_worksheet(sheet)
            copy_sheet.title = sheet_name
            temp_start = sheet_details[sheet_name][0]
            temp_end = sheet_details[sheet_name][1] + 1
            if temp_end < end:
                copy_sheet.delete_rows(temp_end, end + 1 - temp_end)
            if temp_start > start:
                copy_sheet.delete_rows(start, temp_start - start)
        book.remove_sheet(sheet)
        book.save(pathFile.get() + '--copy.xlsx')
        book.close()
    except Exception as e:
        showlog(str(e))
    showlog('finished')


def showlog(text):
    logs.insert(END, str(text) + "\n")
    logs.update()
    print(text)


def selectFile():
    file_path = askopenfilename(filetypes=[('XLSX', '*.xlsx'), ('All Files', '*')])
    pathFile.set(file_path)


def split_sheets_to_files():
    try:
        book = openpyxl.load_workbook(pathFile.get())
        sheets = book.get_sheet_names()
        print(sheets)
        for sheet in sheets:
            # print(sheet.title)
            temps = sheets.copy()
            temps.remove(sheet)
            for t in temps:
                book.remove_sheet(book.get_sheet_by_name(t))
            file_name = pathFile.get() + '--' + sheet + '.xlsx'
            book.save(file_name)
            book = openpyxl.load_workbook(pathFile.get())
        book.close()
    except Exception as e:
        showlog(str(e))
    showlog('finished')


def startIt():
    if len(pathFile.get()) <= 0:
        showinfo('提示', '先选择文件')
        return
    unmerge()


def main():
    logs.grid(row=0, column=0, rowspan=6)

    Button(root, text='选择文件', command=selectFile).grid(row=0, column=1)
    Entry(root, textvariable=pathFile).grid(row=0, column=2)
    Label(root, text='开始行号', ).grid(row=1, column=1)
    Entry(root, textvariable=start_line).grid(row=1, column=2)
    Label(root, text='结束行号', ).grid(row=2, column=1)
    Entry(root, textvariable=end_line).grid(row=2, column=2)
    Label(root, text='哪一列', ).grid(row=3, column=1)
    Entry(root, textvariable=col_line).grid(row=3, column=2)
    Button(root, text='开始', command=startIt).grid(row=4, column=1, columnspan=1)
    Button(root, text='sheets到文件', command=split_sheets_to_files).grid(row=4, column=2, columnspan=1)
    root.mainloop()


root = Tk()
logs = ScrolledText(root, width=40, height=30)
pathFile = StringVar()
start_line = IntVar()
end_line = IntVar()
col_line = IntVar()
startline = -1
if __name__ == '__main__':
    main()
