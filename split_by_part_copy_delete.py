# coding=utf-8
# 复制删除法按地市法人分割汇总文件
import datetime
import os
from tkinter import *
from tkinter.scrolledtext import ScrolledText
from tkinter.ttk import *
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import *
import openpyxl
import _thread

from utils import orgs


def select_file():
    file_path = askopenfilename(filetypes=[('XLSX', '*.xlsx'), ('All Files', '*')])
    pathFile.set(file_path)


def show_log(text):
    logs.insert(END, str(text) + "\n")
    logs.update()
    print(text)


def start_work():
    if len(pathFile.get()) <= 0:
        showinfo('提示', '先选择文件')
        return
    start = start_line.get()
    end = end_line.get()
    column_selected = col_line.get()
    source_file = pathFile.get()
    if video_split_type.get() == 1:
        write_city_file(start, end, column_selected, source_file)
    elif video_split_type.get() == 2:
        write_org_file(start, end, column_selected, source_file)
    return


def write_city_file(start, end, column_selected, source_file):
    results = {}
    try:
        book = openpyxl.load_workbook(source_file)
        sheet = book.active
        for rx in range(start, end + 1):
            org_temp = sheet.cell(row=rx, column=column_selected).value
            if isinstance(org_temp, str):
                org = org_temp.replace("农商", "").replace("商行", "").replace("银", "").replace("行", "") \
                    .replace("市", "").replace(" ", "")
                if len(org) >= 2:
                    if not results.__contains__(org):
                        results[org] = []
                    results[org].append(rx)
        book.close()

        name, suffix = os.path.splitext(pathFile.get())
        file_folder = name + '--' + datetime.datetime.now().strftime('%H%M')
        if not os.path.exists(file_folder):
            os.makedirs(file_folder)
        for city, lines in results.items():
            delete_rows(source_file, start, end, lines, os.sep.join([file_folder, city + suffix]))
            show_log(city + '已完成')
    except Exception as e:
        show_log(str(e))
    show_log('已完成')
    return


def write_org_file(start, end, column_selected, source_file):
    results = {}
    try:
        book = openpyxl.load_workbook(source_file)
        sheet = book.active
        for rx in range(start, end + 1):
            org_temp = sheet.cell(row=rx, column=column_selected).value
            if isinstance(org_temp, str):
                org = org_temp.replace("农商", "").replace("商行", "").replace("银", "").replace("行", "") \
                    .replace("市", "").replace(" ", "")
                if orgs.__contains__(org):
                    city = orgs[org]
                    if not results.__contains__(city):
                        results[city] = {}
                    if not results[city].__contains__(org):
                        results[city][org] = []
                    results[city][org].append(rx)

        name, suffix = os.path.splitext(pathFile.get())
        file_folder = name + '--' + datetime.datetime.now().strftime('%H%M')
        if not os.path.exists(file_folder):
            os.makedirs(file_folder)
        for city, data in results.items():
            city_folder = os.sep.join([file_folder, city])
            if not os.path.exists(city_folder):
                os.makedirs(city_folder)
            for org, lines in data.items():
                delete_rows(source_file, start, end, lines, os.sep.join([city_folder, org + suffix]))
                show_log('\t' + org + '已完成')
            show_log(city + '已完成')
    except Exception as e:
        show_log(str(e))
    show_log('已完成')
    return


def delete_rows(source_file, start, end, lines, file_name):
    book = openpyxl.load_workbook(source_file)
    sheet = book.active
    max_len = len(lines)
    if max_len >= 1:
        temp_start = lines[0]
        temp_end = lines[max_len - 1] + 1
        if temp_end < end:
            sheet.delete_rows(temp_end, end + 1 - temp_end)
        if temp_start > start:
            sheet.delete_rows(start, temp_start - start)
    book.save(file_name)
    book.close()


def begin_work():
    _thread.start_new_thread(start_work, ())


def main():
    logs.grid(row=0, column=0, rowspan=6)

    Button(root, text='选择文件', command=select_file).grid(row=0, column=1)
    Entry(root, textvariable=pathFile).grid(row=0, column=2)
    Label(root, text='开始行号', ).grid(row=1, column=1)
    Entry(root, textvariable=start_line).grid(row=1, column=2)
    Label(root, text='结束行号', ).grid(row=2, column=1)
    Entry(root, textvariable=end_line).grid(row=2, column=2)
    Label(root, text='哪一列', ).grid(row=3, column=1)
    Entry(root, textvariable=col_line).grid(row=3, column=2)
    Radiobutton(root, text=('按地市划分'), variable=video_split_type, value=1).grid(row=4, column=1)
    Radiobutton(root, text=('按法人划分'), variable=video_split_type, value=2).grid(row=4, column=2)

    Button(root, text='开始', command=begin_work).grid(row=5, column=1, columnspan=1)
    Label(root, text='需要另存为xlsx格式\n需按摘取字段排序好\n空行过多需先清除内容后删除', ).grid(row=5, column=2, columnspan=1)
    root.mainloop()


root = Tk()
root.wm_title('按地市/法人分割汇总文件')
logs = ScrolledText(root, width=40, height=30)
pathFile = StringVar()
video_split_type = IntVar(value=1)
start_line = IntVar(value=1)
end_line = IntVar()
col_line = IntVar(value=2)
if __name__ == '__main__':
    main()
