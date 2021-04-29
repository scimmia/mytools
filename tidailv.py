# coding=utf-8
import datetime
from tkinter import *
from tkinter.scrolledtext import ScrolledText
from tkinter.ttk import *
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import *

import openpyxl
import shutil

orgs = {"济南": "济南", "章丘": "济南", "平阴": "济南", "济阳": "济南", "商河": "济南", "莱芜": "济南", "青岛": "青岛", "张店": "淄博", "临淄": "淄博",
        "博山": "淄博", "周村": "淄博", "桓台": "淄博", "高青": "淄博", "沂源": "淄博", "淄川": "淄博", "枣庄": "枣庄", "滕州": "枣庄", "东营": "东营",
        "垦利": "东营", "利津": "东营", "广饶": "东营", "烟台": "烟台", "龙口": "烟台", "莱阳": "烟台", "莱州": "烟台", "蓬莱": "烟台", "招远": "烟台",
        "栖霞": "烟台", "海阳": "烟台", "长岛": "烟台", "潍坊": "潍坊", "青州": "潍坊", "诸城": "潍坊", "寿光": "潍坊", "安丘": "潍坊", "高密": "潍坊",
        "昌邑": "潍坊", "昌乐": "潍坊", "临朐": "潍坊", "济宁": "济宁", "曲阜": "济宁", "兖州": "济宁", "邹城": "济宁", "汶上": "济宁", "泗水": "济宁",
        "微山": "济宁", "鱼台": "济宁", "金乡": "济宁", "嘉祥": "济宁", "梁山": "济宁", "泰山": "泰安", "岱岳": "泰安", "新泰": "泰安", "肥城": "泰安",
        "宁阳": "泰安", "东平": "泰安", "威海": "威海", "荣成": "威海", "文登": "威海", "乳山": "威海", "东港": "日照", "岚山": "日照", "莒县": "日照",
        "五莲": "日照", "滨州": "滨州", "博兴": "滨州", "邹平": "滨州", "惠民": "滨州", "阳信": "滨州", "无棣": "滨州", "德州": "德州", "乐陵": "德州",
        "禹城": "德州", "陵城": "德州", "宁津": "德州", "庆云": "德州", "临邑": "德州", "齐河": "德州", "平原": "德州", "夏津": "德州", "武城": "德州",
        "聊城": "聊城", "临清": "聊城", "高唐": "聊城", "茌平": "聊城", "东阿": "聊城", "阳谷": "聊城", "莘县": "聊城", "润昌": "聊城", "兰山": "临沂",
        "罗庄": "临沂", "河东": "临沂", "沂南": "临沂", "沂水": "临沂", "莒南": "临沂", "临沭": "临沂", "郯城": "临沂", "兰陵": "临沂", "费县": "临沂",
        "平邑": "临沂", "蒙阴": "临沂", "菏泽": "菏泽", "曹县": "菏泽", "定陶": "菏泽", "成武": "菏泽", "单县": "菏泽", "巨野": "菏泽", "郓城": "菏泽",
        "鄄城": "菏泽", "东明": "菏泽"}
cities = {"济南": {"济南": [], "章丘": [], "平阴": [], "济阳": [], "商河": [], "莱芜": []}, "青岛": {"青岛": []},
          "淄博": {"张店": [], "临淄": [], "博山": [], "周村": [], "桓台": [], "高青": [], "沂源": [], "淄川": []},
          "枣庄": {"枣庄": [], "滕州": []}, "东营": {"东营": [], "垦利": [], "利津": [], "广饶": []},
          "烟台": {"烟台": [], "龙口": [], "莱阳": [], "莱州": [], "蓬莱": [], "招远": [], "栖霞": [], "海阳": [], "长岛": []},
          "潍坊": {"潍坊": [], "青州": [], "诸城": [], "寿光": [], "安丘": [], "高密": [], "昌邑": [], "昌乐": [], "临朐": []},
          "济宁": {"济宁": [], "曲阜": [], "兖州": [], "邹城": [], "汶上": [], "泗水": [], "微山": [], "鱼台": [], "金乡": [], "嘉祥": [],
                 "梁山": []}, "泰安": {"泰山": [], "岱岳": [], "新泰": [], "肥城": [], "宁阳": [], "东平": []},
          "威海": {"威海": [], "荣成": [], "文登": [], "乳山": []}, "日照": {"东港": [], "岚山": [], "莒县": [], "五莲": []},
          "滨州": {"滨州": [], "博兴": [], "邹平": [], "惠民": [], "阳信": [], "无棣": []},
          "德州": {"德州": [], "乐陵": [], "禹城": [], "陵城": [], "宁津": [], "庆云": [], "临邑": [], "齐河": [], "平原": [], "夏津": [],
                 "武城": []}, "聊城": {"聊城": [], "临清": [], "高唐": [], "茌平": [], "东阿": [], "阳谷": [], "莘县": [], "润昌": []},
          "临沂": {"兰山": [], "罗庄": [], "河东": [], "沂南": [], "沂水": [], "莒南": [], "临沭": [], "郯城": [], "兰陵": [], "费县": [],
                 "平邑": [], "蒙阴": []},
          "菏泽": {"菏泽": [], "曹县": [], "定陶": [], "成武": [], "单县": [], "巨野": [], "郓城": [], "鄄城": [], "东明": []}}


def add_to_results(the_results, org, part, row):
    if not orgs.__contains__(org):
        showlog('错误：' + org + '\t' + row)
        return -1
    else:
        if not the_results.__contains__(org):
            the_results[org] = {}
        if not the_results[org].__contains__(part):
            the_results[org][part] = {}
            the_results[org][part]['eight'] = []
            the_results[org][part]['nine'] = []
        else:
            showlog('重复p----' + part)
        the_results[org][part]['row'] = row + 1


def read_sheet_data(book, sheet_name, the_results):
    all_orgs = orgs.keys()
    src_sheet = book.get_sheet_by_name(sheet_name)
    for index, row in enumerate(src_sheet.values):
        if row[0][4:6] in all_orgs:
            org = row[0][4:6]
        elif row[0][2:4] in all_orgs:
            org = row[0][2:4]
        elif row[0][0:2] in all_orgs:
            org = row[0][0:2]
        else:
            org = None
        if the_results.__contains__(org):
            for part, value in the_results[org].items():
                if part.endswith('支行') or part.endswith('分理处') or (part.endswith('营业部') and part != '营业部'):
                    temp = part.replace('支行', '').replace('分理处', '').replace('营业部','')
                else:
                    temp = part
                if row[0].find(temp) >= 0:
                    the_results[org][part][sheet_name].append((row[0], row[1]))

                    # print('%s--%d--%s--%d' % (part, value,row[0], row[1]))
                    # to_sheet['%s%d' % (col_name, value + 1)] = row[1]
                    break


def from_sheet_to_sheet(src_sheet, to_sheet, col_name, the_results):
    all_orgs = orgs.keys()
    for index, row in enumerate(src_sheet.values):
        if row[0][4:6] in all_orgs:
            org = row[0][4:6]
        elif row[0][2:4] in all_orgs:
            org = row[0][2:4]
        elif row[0][0:2] in all_orgs:
            org = row[0][0:2]
        else:
            org = None
        if the_results.__contains__(org):
            for part, value in the_results[org].items():
                if part.endswith('支行') or part.endswith('分理处'):
                    temp = part.replace('支行', '').replace('分理处', '')
                else:
                    temp = part
                if row[0].find(temp) >= 0:
                    the_results[org][part][src_sheet.title].append((row[0], row[1]))

                    # print('%s--%d--%s--%d' % (part, value,row[0], row[1]))
                    # to_sheet['%s%d' % (col_name, value + 1)] = row[1]
                    break


# 先将取消合并的单元格，再操作。
def doIt():
    try:
        the_results = {}
        book = openpyxl.load_workbook(pathFile.get())
        sheet = book.get_sheet_by_name('示范单位')
        for index, row in enumerate(sheet.values):
            print(row[2] + '---' + row[4] + '----')
            org = row[2].replace("农商", "").replace("商行", "").replace("银", "").replace("行", "")
            part = row[4]
            # if part.endswith('营业部') and part != '营业部':
            #     part = part.replace('营业部','')
            print('---' + org + '----' + part)

            add_to_results(the_results, org, part, index)
        read_sheet_data(book, 'eight', the_results)
        read_sheet_data(book, 'nine', the_results)

        none_eight = []
        none_nine = []
        for city, city_value in the_results.items():
            for part, value in city_value.items():
                row = value['row']
                eight = value['eight']
                if len(eight) == 0:
                    # showlog('%s--%d---八月无数据' % (part, row))
                    none_eight.append((part,row))
                elif len(eight) == 1:
                    sheet['%s%d' % ('F', row)] = eight[0][1]
                else:
                    used_data = False
                    for data in eight:
                        if data[0].endswith(part):
                            sheet['%s%d' % ('F', row)] = data[1]
                            used_data = True
                            showlog('%s--%s--%s--%f---采用了' % (city,part,data[0], data[1]))
                            break
                    if not used_data:
                        showlog('%s--%s--%d---八月多数据' % (city, part, row))
                    print(eight)
                nine = value['nine']
                if len(nine) == 0:
                    # showlog('%s--%d---八月无数据' % (part, row))
                    none_nine.append((part,row))
                elif len(nine) == 1:
                    sheet['%s%d' % ('G', row)] = nine[0][1]
                else:
                    used_data = False
                    for data in nine:
                        if data[0].endswith(part):
                            sheet['%s%d' % ('G', row)] = data[1]
                            used_data = True
                            showlog('%s--%s--%s--%f---采用了' % (city,part,data[0], data[1]))
                            break
                    if not used_data:
                        showlog('%s--%s--%d---九月多数据' % (city, part, row))
                    print(nine)

                # if len(nine) == 0:
                #     showlog('%s--%d---九月无数据' % (part, row))
                # elif len(nine) == 1:
                #     sheet['%s%d' % ('G', row)] = nine[0][1]
                # else:
                #     showlog('%s--%d---九月多数据' % (part, row))
                #     print(nine)

                # if len(value['eight']) == 1 and len(value['nine']) == 1:
                #     sheet['%s%d' % ('F', value['row'])] = value['eight'][0]
                #     sheet['%s%d' % ('G', value['row'])] = value['nine'][0]
                # else:
                #     if len(value['eight']) == 0:
                #         showlog('%s--八月无数据' % part)
                #     if len(value['nine']) == 0:
                #         showlog('%s--九月无数据' % part)
        for e in none_eight:
            showlog('%s--%d---八月无数据' % e)
        for e in none_nine:
            showlog('%s--%d---九月无数据' % e)
        # eight = book.get_sheet_by_name('8')
        # from_sheet_to_sheet(eight, sheet, 'F', the_results)
        # nine = book.get_sheet_by_name('9')
        # from_sheet_to_sheet(nine, sheet, 'G', the_results)
        book.save(pathFile.get() + '--copy.xlsx')
        print('')
    except Exception as e:
        showlog(str(e))
    showlog('finished')


def showlog(text):
    logs.insert(END, str(text) + "\n")
    logs.update()
    print(text)


def selectFile():
    file_path = askopenfilename(filetypes=[('XLSX', '*.xlsx'), ('All Files', '*')])
    pathFile.set(file_path)


def select_src_file():
    file_path = askopenfilename(filetypes=[('XLSX', '*.xlsx'), ('All Files', '*')])
    src_pathFile.set(file_path)


def split_sheets_to_files():
    try:
        book = openpyxl.load_workbook(pathFile.get())
        sheets = book.get_sheet_names()
        print(sheets)
        for sheet in sheets:
            # print(sheet.title)
            temps = sheets.copy()
            temps.remove(sheet)
            for t in temps:
                book.remove_sheet(book.get_sheet_by_name(t))
            file_name = pathFile.get() + '--' + sheet + '.xlsx'
            book.save(file_name)
            book = openpyxl.load_workbook(pathFile.get())
        book.close()
    except Exception as e:
        showlog(str(e))
    showlog('finished')


def startIt():
    if len(pathFile.get()) <= 0:
        showinfo('提示', '先选择源文件')
        return
    # if len(src_pathFile.get()) <= 0:
    #     showinfo('提示', '先选择数据文件')
    #     return
    doIt()


def main():
    logs.grid(row=0, column=0, rowspan=6)

    Button(root, text='选择源文件', command=selectFile).grid(row=0, column=1)
    Entry(root, textvariable=pathFile).grid(row=0, column=2)
    # Button(root, text='选择数据文件', command=select_src_file()).grid(row=1, column=1)
    # Entry(root, textvariable=pathFile).grid(row=1, column=2)
    Label(root, text='开始行号', ).grid(row=1, column=1)
    Entry(root, textvariable=start_line).grid(row=1, column=2)
    Label(root, text='结束行号', ).grid(row=2, column=1)
    Entry(root, textvariable=end_line).grid(row=2, column=2)
    Label(root, text='哪一列', ).grid(row=3, column=1)
    Entry(root, textvariable=col_line).grid(row=3, column=2)
    Button(root, text='开始', command=startIt).grid(row=4, column=1, columnspan=1)
    Button(root, text='sheets到文件', command=split_sheets_to_files).grid(row=4, column=2, columnspan=1)
    root.mainloop()


root = Tk()
logs = ScrolledText(root, width=40, height=30)
pathFile = StringVar()
src_pathFile = StringVar()
start_line = IntVar()
end_line = IntVar()
col_line = IntVar()
startline = -1
if __name__ == '__main__':
    main()
